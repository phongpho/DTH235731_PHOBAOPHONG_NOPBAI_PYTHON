from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "nhanvien.xlsx"

# ---------------------- Lớp đối tượng ----------------------
class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = int(tuoi)

# ---------------------- Ghi file Excel ----------------------
def ghi_file_excel(ds_nv):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"

    # Ghi tiêu đề
    ws.append(["STT", "Mã", "Tên", "Tuổi"])

    # Ghi dữ liệu nhân viên
    for i, nv in enumerate(ds_nv, start=1):
        ws.append([i, nv.ma, nv.ten, nv.tuoi])

    wb.save(FILE_NAME)
    print("Đã lưu danh sách nhân viên vào file", FILE_NAME)

# ---------------------- Đọc file Excel ----------------------
def doc_file_excel():
    ds_nv = []
    if not os.path.exists(FILE_NAME):
        print("Chưa có file dữ liệu.")
        return ds_nv

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # Bỏ qua dòng tiêu đề (dòng 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2] and row[3]:
            ds_nv.append(NhanVien(row[1], row[2], row[3]))

    print("Đã đọc dữ liệu từ file Excel.")
    return ds_nv

# ---------------------- Hiển thị danh sách ----------------------
def hien_thi(ds_nv):
    if not ds_nv:
        print("Danh sách trống.")
        return
    print("{:<5}{:<10}{:<15}{:<5}".format("STT", "Mã", "Tên", "Tuổi"))
    for i, nv in enumerate(ds_nv, start=1):
        print("{:<5}{:<10}{:<15}{:<5}".format(i, nv.ma, nv.ten, nv.tuoi))

# ---------------------- Thêm nhân viên ----------------------
def them_nhan_vien(ds_nv):
    ma = input("Nhập mã NV: ").strip()
    ten = input("Nhập tên NV: ").strip()
    tuoi = int(input("Nhập tuổi: "))
    ds_nv.append(NhanVien(ma, ten, tuoi))
    print("Đã thêm nhân viên thành công.")

# ---------------------- Sắp xếp ----------------------
def sap_xep(ds_nv):
    ds_nv.sort(key=lambda nv: nv.tuoi)
    print("Đã sắp xếp nhân viên theo tuổi tăng dần.")

# ---------------------- MENU CHÍNH ----------------------
def main():
    ds_nv = []

    while True:
        print("\n=== QUẢN LÝ NHÂN VIÊN (Excel) ===")
        print("1. Xem danh sách")
        print("2. Thêm nhân viên")
        print("3. Sắp xếp theo tuổi")
        print("4. Lưu vào file Excel")
        print("5. Đọc từ file Excel")
        print("6. Thoát")
        chon = input("Chọn chức năng (1-6): ")

        match chon:
            case "1":
                hien_thi(ds_nv)
            case "2":
                them_nhan_vien(ds_nv)
            case "3":
                sap_xep(ds_nv)
            case "4":
                ghi_file_excel(ds_nv)
            case "5":
                ds_nv = doc_file_excel()
            case "6":
                print("Tạm biệt!")
                break
            case _:
                print("Lựa chọn không hợp lệ!")

if __name__ == "__main__":
    main()